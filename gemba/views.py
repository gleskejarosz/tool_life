import csv
import os

import xlwt
from django.core.paginator import Paginator
from xlwt import XFStyle, Font
from datetime import datetime, timezone, timedelta

from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import HttpResponse
from django.shortcuts import render, get_object_or_404, redirect
from django.urls import reverse_lazy
from django.views import View
from django.views.generic import TemplateView, UpdateView, DeleteView, DetailView, ListView
from django.db.models import Q

from gemba.forms import ParetoDetailForm, DowntimeMinutes, ScrapQuantity, NewPareto, ParetoUpdateForm, \
    NotScheduledToRunUpdateForm, ParetoTotalQtyDetailForm, ParetoDetailUpdateForm
from gemba.models import Pareto, ParetoDetail, DowntimeModel, DowntimeDetail, ScrapModel, ScrapDetail, DowntimeUser, \
    DowntimeGroup, ScrapUser, LineHourModel, JobModel2, SHIFT_CHOICES, TC, HC, Editors, AM, PM, NS, LineUser
from tools.views import tools_update


class GembaIndex(TemplateView):
    template_name = 'gemba/index.html'


def downtimes_view(request):
    downtimes = DowntimeDetail.objects.all().order_by("-modified")
    last_week_downtimes = DowntimeDetail.objects.filter(pareto_date__gte=datetime.now() - timedelta(days=7)).order_by(
        "-modified")
    paginator = Paginator(last_week_downtimes, 100)

    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    return render(request,
                  template_name="gemba/downtimes_view.html",
                  context={
                      "filter": downtimes,
                      "page_obj": page_obj,
                  })


def scraps_view(request):
    scraps = ScrapDetail.objects.all().order_by("-modified")
    last_week_scraps = ScrapDetail.objects.filter(pareto_date__gte=datetime.now() - timedelta(days=7)).order_by(
        "-modified")
    paginator = Paginator(last_week_scraps, 100)

    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request,
                  template_name="gemba/scraps_view.html",
                  context={
                      "filter": scraps,
                      "page_obj": page_obj,
                  })


def pareto_details_view(request):
    pareto_details = ParetoDetail.objects.all().order_by("-modified")
    last_week_pareto_details = ParetoDetail.objects.filter(pareto_date__gte=datetime.now() - timedelta(days=7)
                                                           ).order_by("-modified")
    paginator = Paginator(last_week_pareto_details, 100)

    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request,
                  template_name="gemba/pareto_details_view.html",
                  context={
                      "filter": pareto_details,
                      "page_obj": page_obj,
                  })


@login_required
def downtime_detail_create(request, pk):
    downtime = get_object_or_404(DowntimeModel, pk=pk)
    pareto_qs = Pareto.objects.filter(user=request.user, completed=False)
    pareto = pareto_qs[0]
    pareto_id = pareto.id
    pareto_date = pareto.pareto_date
    job = pareto.job_otg
    line = pareto.line

    if job is None:
        return redirect("gemba_app:pareto-summary")

    # downtime_qs = DowntimeDetail.objects.filter(user=request.user, completed=False, downtime_id=pk, job=job)
    #
    # if downtime_qs.exists():
    #     downtime_id = downtime_qs[0].id
    #     downtime = DowntimeDetail.objects.get(id=downtime_id)
    #     old_time = downtime.minutes
    #     form = DowntimeMinutes(request.POST or None)
    #     if form.is_valid():
    #         minutes = form.cleaned_data["minutes"]
    #         new_time = minutes + old_time
    #         downtime.minutes = new_time
    #         downtime.frequency += 1
    #         downtime.save()
    #         return redirect("gemba_app:pareto-summary")
    # else:
    form = DowntimeMinutes(request.POST or None)

    if form.is_valid():
        minutes = form.cleaned_data["minutes"]
        user = request.user
        downtime_elem = DowntimeDetail.objects.create(downtime=downtime, minutes=minutes, user=user, job=job,
                                                      pareto_id=pareto_id, pareto_date=pareto_date, line=line)
        pareto.downtimes.add(downtime_elem)
        return redirect("gemba_app:pareto-summary")

    return render(
        request,
        template_name="form.html",
        context={"form": form}
    )


@login_required
def scrap_detail_create(request, pk):
    scrap = get_object_or_404(ScrapModel, pk=pk)
    pareto_qs = Pareto.objects.filter(user=request.user, completed=False)
    pareto = pareto_qs[0]
    pareto_id = pareto.id
    pareto_date = pareto.pareto_date
    job = pareto.job_otg
    line = pareto.line

    if job is None:
        return redirect("gemba_app:pareto-summary")

    form = ScrapQuantity(request.POST or None)
    if form.is_valid():
        qty = form.cleaned_data["qty"]
        user = request.user
        scrap_qs = ScrapDetail.objects.filter(user=request.user, completed=False, scrap=scrap, job=job)

        if scrap_qs.exists():
            scrap_elem = ScrapDetail.objects.get(scrap=scrap, user=user, job=job)
            scrap_elem.qty += qty
            scrap_elem.save()
            return redirect("gemba_app:pareto-summary")
        else:
            scrap_elem = ScrapDetail.objects.create(scrap=scrap, qty=qty, user=user, job=job,
                                                    pareto_id=pareto_id, pareto_date=pareto_date, line=line)
            pareto.scrap.add(scrap_elem)
            return redirect("gemba_app:pareto-summary")

    return render(
        request,
        template_name="form.html",
        context={"form": form}
    )


@login_required
def pareto_detail_create(request):
    user = request.user
    pareto_qs = Pareto.objects.filter(user=user, completed=False)
    pareto = pareto_qs[0]
    job = pareto.job_otg
    line = pareto.line
    if job is None:
        return redirect("gemba_app:pareto-summary")

    job_id = pareto.job_otg_id
    job_obj = JobModel2.objects.get(id=job_id)
    target = job_obj.target
    takt_time = round(60 / target, ndigits=5)
    pareto_details_qs = ParetoDetail.objects.filter(user=user, completed=False, job=job)

    total_output = 0
    total_good = 0

    for elem in pareto_details_qs:
        total_output += elem.output
        total_good += elem.good

    pareto_id = pareto.id
    pareto_date = pareto.pareto_date
    down_group = DowntimeGroup.objects.get(user=user)
    calc_option = down_group.calculation

    if calc_option == TC:
        form = ParetoTotalQtyDetailForm(request.POST or None)
        if form.is_valid():
            output = form.cleaned_data["output"]
            good = form.cleaned_data["good"]

            new_output = output - total_output
            new_good = good - total_good
            scrap = output - good

            if pareto_details_qs.exists():
                pareto_elem = ParetoDetail.objects.get(user=user, job=job, pareto_id=pareto_id)
                modified = pareto_elem.modified
                pareto_elem.output += new_output
                pareto_elem.good += new_good
                pareto_elem.scrap = scrap
                pareto_elem.save()

                # tools update
                tools_update(job=job, output=new_output, target=target, modified=modified)

                return redirect("gemba_app:pareto-summary")

            else:
                pareto_item = ParetoDetail.objects.create(job=job, output=output, user=user, good=good,
                                                          pareto_id=pareto_id, line=line,
                                                          pareto_date=pareto_date, scrap=scrap, takt_time=takt_time)
                modified = pareto_item.modified

                pareto.jobs.add(pareto_item)

                # tools update
                tools_update(job=job, output=output, target=target, modified=modified)

                return redirect("gemba_app:pareto-summary")

        return render(
            request,
            template_name="gemba/form.html",
            context={
                "form": form,
                "total_output": total_output,
                "total_good": total_good,
            })
    else:
        form = ParetoDetailForm(request.POST or None)
        if form.is_valid():
            good = form.cleaned_data["good"]

            scrap_qs = ScrapDetail.objects.filter(user=user, completed=False, pareto_id=pareto_id, job=job)
            scrap = 0
            for scrap_elem in scrap_qs:
                scrap += scrap_elem.qty

            good_qs = ParetoDetail.objects.filter(user=user, completed=False, pareto_id=pareto_id, job=job)
            old_good = 0
            for good_elem in good_qs:
                old_good += good_elem.good

            job_elem = JobModel2.objects.get(name=job)
            inner_size = job_elem.inner_size

            cal_good = good * inner_size
            output = old_good + cal_good + scrap

            if pareto_details_qs.exists():
                pareto_elem = ParetoDetail.objects.get(user=user, job=job, pareto_id=pareto_id)
                pareto_elem.output = output
                pareto_elem.good += cal_good
                pareto_elem.scrap = scrap
                pareto_elem.save()
                return redirect("gemba_app:pareto-summary")
            else:
                pareto_item = ParetoDetail.objects.create(job=job, output=output, user=user, good=cal_good,
                                                          pareto_id=pareto_id, line=line,
                                                          pareto_date=pareto_date, scrap=scrap, takt_time=takt_time)
                pareto.jobs.add(pareto_item)
                return redirect("gemba_app:pareto-summary")

        return render(
            request,
            template_name="form.html",
            context={
                "form": form,
                "total_output": total_output,
                "total_good": total_good,
            })


class ParetoSummary(LoginRequiredMixin, View):
    def get(self, *args, **kwargs):
        user = self.request.user
        pareto_qs = Pareto.objects.filter(user=user, completed=False)
        if pareto_qs.exists():
            pareto = pareto_qs[0]

            pareto_id = pareto.id
            pareto_status = ""
            status = pareto.completed
            hours = pareto.hours
            time_stamp = pareto.time_stamp
            not_scheduled_to_run = pareto.not_scheduled_to_run

            message_status = ""
            job_otg = pareto.job_otg
            if job_otg is None:
                message_status = "Display"

            available_time = available_time_cal(status=status, hours=hours, time_stamp=time_stamp,
                                                not_scheduled_to_run=not_scheduled_to_run)

            pareto_details = ParetoDetail.objects.filter(user=user, completed=False, pareto_id=pareto_id)
            total_good = 0
            total_output = 0
            if pareto_details.exists():
                for scrap_cal in pareto_details:
                    output = scrap_cal.output
                    total_output += output
                    good = scrap_cal.good
                    total_good += good
            total_scrap_cal = total_output - total_good

            quality = quality_cal(good=total_good, output=total_output)

            down_qs = DowntimeDetail.objects.filter(user=user, completed=False, pareto_id=pareto_id).order_by(
                "modified")
            total_down = 0
            if down_qs.exists():
                for down in down_qs:
                    quantity = down.minutes
                    total_down += quantity

            availability = availability_cal(available_time=available_time, downtime=total_down)

            scrap_qs = ScrapDetail.objects.filter(user=user, completed=False, pareto_id=pareto_id).order_by("modified")
            total_scrap = 0
            for scrap_elem in scrap_qs:
                qty = scrap_elem.qty
                total_scrap += qty

            performance_numerator = 0

            if pareto_details.exists():
                for pareto_detail in pareto_details:
                    job_elem = pareto_detail.job_id
                    job_model = JobModel2.objects.get(id=job_elem)
                    takt_time = round(60 / job_model.target, 5)
                    qty_elem = pareto_detail.output
                    performance_numerator += (qty_elem * takt_time)

            performance = performance_cal(performance_numerator=performance_numerator, available_time=available_time,
                                          downtime=total_down)

            oee = oee_cal(availability=availability, performance=performance, quality=quality)

            return render(self.request,
                          template_name='gemba/pareto.html',
                          context={
                              "pareto_list": pareto,
                              "down_qs": down_qs,
                              "scrap_qs": scrap_qs,
                              "user": user,
                              "pareto_status": pareto_status,
                              "message_status": message_status,
                              "available_time": available_time,
                              "availability": availability,
                              "quality": quality,
                              "performance": performance,
                              "oee": oee,
                              "total_scrap": total_scrap,
                              "total_down": total_down,
                              "total_output": total_output,
                              "total_good": total_good,
                              "total_scrap_cal": total_scrap_cal,
                          },
                          )
        else:
            pareto_status = "Not exist"
            message_status = ""
            user = self.request.user
            available_time = 0
            availability = 0
            quality = 0
            performance = 0
            oee = 0
            total_scrap = 0
            total_down = 0
            total_output = 0
            total_good = 0
            total_scrap_cal = 0

            return render(self.request,
                          template_name='gemba/pareto.html',
                          context={
                              "available_time": available_time,
                              "user": user,
                              "availability": availability,
                              "quality": quality,
                              "performance": performance,
                              "oee": oee,
                              "pareto_status": pareto_status,
                              "total_scrap": total_scrap,
                              "total_down": total_down,
                              "total_output": total_output,
                              "total_good": total_good,
                              "total_scrap_cal": total_scrap_cal,
                              "message_status": message_status,
                          }
                          )


def available_time_cal(status, hours, time_stamp, not_scheduled_to_run):
    if status is True:
        available_time = int(hours) * 60 - not_scheduled_to_run
    else:
        now = datetime.now()
        sec_now = int(now.strftime('%S'))
        sec_now += int(now.strftime('%M')) * 60
        sec_now += int(now.strftime('%H')) * 60 * 60

        sec_start = int(time_stamp.strftime('%S'))
        sec_start += int(time_stamp.strftime('%M')) * 60
        sec_start += int(time_stamp.strftime('%H')) * 60 * 60

        available_time = round((sec_now - sec_start) / 60.0)
    return available_time


def performance_cal(performance_numerator, available_time, downtime):
    if available_time - downtime != 0:
        performance = round((performance_numerator / (available_time - downtime)) * 100, ndigits=2)
    else:
        performance = 0
    return performance


def quality_cal(good, output):
    if output != 0:
        quality = round((good / output) * 100, ndigits=2)
    else:
        quality = 0
    return quality


def availability_cal(available_time, downtime):
    if available_time == 0:
        availability = 0
    else:
        availability = round(((available_time - downtime) / available_time * 100), ndigits=2)
    return availability


def oee_cal(availability, performance, quality):
    divider = 10000

    if availability == 0 or performance == 0 or quality == 0:
        return 0

    availability = int(availability * 100)
    divider *= 100
    performance = int(performance * 100)
    divider *= 100
    quality = int(quality * 100)
    divider *= 100

    oee = round(availability * performance * quality / divider, ndigits=2)

    return oee


def pareto_detail_view(request, pk):
    pareto = Pareto.objects.get(pk=pk)
    report_list = oee_calculation(pareto)
    return render(request,
                  template_name='gemba/pareto_details.html',
                  context={
                      "pareto_list": pareto,
                      "report_list": report_list,
                  },
                  )


# when depends on the status
def oee_calculation(pareto):
    hours = pareto.hours
    not_scheduled_to_run = pareto.not_scheduled_to_run
    status = pareto.completed
    time_stamp = pareto.time_stamp

    available_time = available_time_cal(status=status, hours=hours, time_stamp=time_stamp,
                                        not_scheduled_to_run=not_scheduled_to_run)
    calculation = {}

    output = 0
    good = 0
    performance_numerator = 0

    calculation["available_time"] = available_time

    for detail in pareto.jobs.all():
        output += detail.output
        good += detail.good
        performance_numerator += (detail.output * detail.takt_time)

    if status is True:
        quality = pareto.quality
    else:
        quality = quality_cal(good=good, output=output)
    calculation["quality"] = quality
    calculation["good"] = good

    downtime = 0
    for down_elem in pareto.downtimes.all():
        downtime += down_elem.minutes

    calculation["output"] = output
    calculation["downtime"] = downtime

    if status is True:
        performance = pareto.performance
    else:
        performance = performance_cal(performance_numerator=performance_numerator, available_time=available_time,
                                      downtime=downtime)
    calculation["performance"] = performance

    availability = availability_cal(available_time=available_time, downtime=downtime)
    calculation["availability"] = availability

    if status is True:
        oee = pareto.oee
    else:
        oee = oee_cal(availability=availability, performance=performance, quality=quality)

    calculation["oee"] = oee
    scrap = 0
    for scrap_elem in pareto.scrap.all():
        scrap += scrap_elem.qty

    calculation["scrap"] = scrap

    return calculation


# when status not completed, but wants to see how it will be saved
def final_oee_calculation(pareto):
    hours = pareto.hours
    not_scheduled_to_run = pareto.not_scheduled_to_run

    available_time = int(hours) * 60 - not_scheduled_to_run

    calculation = {}

    output = 0
    good = 0
    performance_numerator = 0

    calculation["available_time"] = available_time

    for detail in pareto.jobs.all():
        output += detail.output
        good += detail.good
        performance_numerator += (detail.output * detail.takt_time)

    quality = quality_cal(good=good, output=output)
    calculation["quality"] = quality
    calculation["good"] = good

    downtime = 0
    for down_elem in pareto.downtimes.all():
        downtime += down_elem.minutes

    calculation["output"] = output
    calculation["downtime"] = downtime

    performance = performance_cal(performance_numerator=performance_numerator, available_time=available_time,
                                  downtime=downtime)
    calculation["performance"] = performance

    availability = availability_cal(available_time=available_time, downtime=downtime)
    calculation["availability"] = availability

    oee = oee_cal(availability=availability, performance=performance, quality=quality)

    calculation["oee"] = oee
    scrap = 0
    for scrap_elem in pareto.scrap.all():
        scrap += scrap_elem.qty

    calculation["scrap"] = scrap

    return calculation


def before_close_pareto(request):
    pareto = Pareto.objects.get(user=request.user, completed=False)

    calculation = final_oee_calculation(pareto)

    available_time = calculation["available_time"]
    output = calculation["output"]
    good = calculation["good"]
    downtime = calculation["downtime"]
    availability = calculation["availability"]
    quality = calculation["quality"]
    performance = calculation["performance"]
    oee = calculation["oee"]
    scrap = calculation["scrap"]
    scrap_compare = output - good - scrap

    return render(request,
                  template_name='gemba/pareto_before_close.html',
                  context={
                      "pareto_list": pareto,
                      "available_time": available_time,
                      "output": output,
                      "good": good,
                      "downtime": downtime,
                      "availability": availability,
                      "quality": quality,
                      "performance": performance,
                      "oee": oee,
                      "scrap": scrap,
                      "scrap_compare": scrap_compare,
                  },
                  )


@login_required
def final_confirmation_before_close_pareto(request):
    return render(request,
                  template_name='gemba/pareto_final_qs_before_close.html')


@login_required
def close_pareto(request):
    pareto = Pareto.objects.get(user=request.user, completed=False)

    calculation = final_oee_calculation(pareto)

    availability = calculation["availability"]
    quality = calculation["quality"]
    performance = calculation["performance"]
    oee = calculation["oee"]

    down_items = pareto.downtimes.all()
    down_items.update(completed=True)
    for item in down_items:
        item.save()

    scrap_items = pareto.scrap.all()
    scrap_items.update(completed=True)
    for item in scrap_items:
        item.save()

    details_item = pareto.jobs.all()
    details_item.update(completed=True)
    for item in details_item:
        item.save()

    pareto.availability = availability
    pareto.performance = performance
    pareto.quality = quality
    pareto.oee = oee
    pareto.completed = True
    pareto.save()

    return redirect("gemba_app:index")


# more than one pareto
def get_details_to_display(object_list):
    """
    Organizing data to be displayed in the form of a list of dictionaries, taking into account OEE calculations.
    Used in Daily OEE Report.
    """
    report_list = []
    for pareto in object_list:
        date = pareto.pareto_date
        id = pareto.id
        shift = pareto.shift
        user = pareto.user.username
        line = pareto.line
        status = pareto.completed
        hours = pareto.hours
        time_stamp = pareto.time_stamp
        job_otg = pareto.job_otg
        not_scheduled_to_run = pareto.not_scheduled_to_run
        available_time = available_time_cal(status=status, hours=hours, time_stamp=time_stamp,
                                            not_scheduled_to_run=not_scheduled_to_run)

        output = 0
        good = 0
        performance_numerator = 0

        for detail in pareto.jobs.all():
            output += detail.output
            good += detail.good
            performance_numerator += (detail.output * detail.takt_time)

        if status is True:
            quality = pareto.quality
        else:
            quality = quality_cal(good=good, output=output)

        downtime = 0
        for down_elem in pareto.downtimes.all():
            downtime += down_elem.minutes

        if status is True:
            performance = pareto.performance
        else:
            performance = performance_cal(performance_numerator=performance_numerator, available_time=available_time,
                                          downtime=downtime)

        if status is True:
            availability = pareto.availability
        else:
            availability = availability_cal(available_time=available_time, downtime=downtime)

        if status is True:
            oee = pareto.oee
        else:
            oee = oee_cal(availability=availability, performance=performance, quality=quality)

        scrap = 0
        for scrap_elem in pareto.scrap.all():
            scrap += scrap_elem.qty

        report_list.append({
            "id": id,
            "date": date,
            "shift": shift,
            "job_otg": job_otg,
            "user": user,
            "line": line,
            "availability": availability,
            "performance": performance,
            "quality": quality,
            "oee": oee,
        })
    return report_list


class DailyParetoSearchResultsView(ListView):
    model = Pareto
    template_name = "gemba/daily_oee_report.html"

    def get_queryset(self):
        query = self.request.GET.get("q")
        report_list = Pareto.objects.filter(
            Q(pareto_date__exact=query)
        ).order_by("-user", "shift")
        object_list = get_details_to_display(object_list=report_list)
        return object_list


@login_required
def tableau_export(request, pk):
    pareto = Pareto.objects.get(pk=pk)

    date = pareto.pareto_date
    shift = pareto.shift
    if shift == AM:
        shift = "AM"
    elif shift == PM:
        shift = "PM"
    elif shift == NS:
        shift = "NS"

    user = pareto.user
    hours = pareto.hours
    not_scheduled_to_run = pareto.not_scheduled_to_run
    available_time = int(hours) * 60 - not_scheduled_to_run
    performance = pareto.performance / 100
    quality = pareto.quality / 100
    availability = pareto.availability / 100
    oee = pareto.oee / 100
    ops = pareto.ops

    response = HttpResponse(content_type="application/ms-excel")
    filename = f"{date} - {shift} - {user.username}.xls"
    # dest = ('\\').join(source.split('\\')[:1})
    response["Content-Disposition"] = 'attachment; filename= "{}"'.format(filename)

    wb = xlwt.Workbook(encoding="utf-8")
    ws = wb.add_sheet("Report")

    style3 = XFStyle()
    style3.num_format_str = "YYYY-MM-DD"
    # A1 Date
    ws.write(0, 0, date, style3)
    # A2 Line name <- fix it later
    ws.write(1, 0, user.username)
    # A3 Shift
    ws.write(2, 0, shift)
    # A4 Available time
    ws.write(3, 0, available_time)
    # H1 Operators
    ws.write(0, 7, ops)

    style2 = XFStyle()
    style2.num_format_str = "0%"
    # O1 Availability
    ws.write(0, 14, "Availability")
    ws.write(0, 15, availability, style2)
    # O2 Performance
    ws.write(1, 14, "Performance")
    ws.write(1, 15, performance, style2)
    # 03 Quality
    ws.write(2, 14, "Quality")
    ws.write(2, 15, quality, style2)
    # 04 OEE
    ws.write(3, 14, "OEE")
    ws.write(3, 15, oee, style2)

    jobs = []
    col = 1
    for elem in pareto.jobs.all():
        elem_job = elem.job.name
        jobs.append(elem_job)
        elem_job_id = elem.job.id
        job_obj = JobModel2.objects.get(id=elem_job_id)
        target = job_obj.target
        elem_takt_time = round(60 / target, ndigits=4)
        elem_output = elem.output
        elem_good = elem.good
        ws.write(0, col, elem_job)
        ws.write(1, col, elem_takt_time)
        ws.write(2, col, elem_output)
        ws.write(3, col, elem_good)
        col += 1

    group = DowntimeGroup.objects.get(user=user)
    scrap_qs = ScrapUser.objects.filter(group=group).order_by("gemba")

    vector = 4  # start scrap reasons from row 5

    for idx, scrap_obj in enumerate(scrap_qs):
        scrap_id = scrap_obj.scrap.id
        scrap_item = ScrapModel.objects.get(id=scrap_id)
        scrap_name = scrap_item.description
        ws.write(vector + idx, 0, scrap_name)

        scraps_exist_qs = ScrapDetail.objects.filter(pareto_id=pk).filter(scrap=scrap_id)
        for elem in scraps_exist_qs:
            job = elem.job.name
            column = jobs.index(job) + 1
            qty = elem.qty
            ws.write(vector + idx, column, qty)

    down_qs = DowntimeUser.objects.filter(group=group).order_by("gemba")

    for idx, down_obj in enumerate(down_qs):
        down_id = down_obj.downtime.id
        down_item = DowntimeModel.objects.get(id=down_id)
        down_code = down_item.code
        down_name = down_item.description
        ws.write(idx, 9, down_code)
        ws.write(idx, 10, down_name)

        down_exist_obj = DowntimeDetail.objects.filter(pareto_id=pk).filter(downtime=down_id)
        minutes = 0
        frequency = 0
        if down_exist_obj.exists():
            for down_obj in down_exist_obj:
                obj_min = down_obj.minutes
                obj_freq = down_obj.frequency
                minutes += obj_min
                frequency += obj_freq
            ws.write(idx, 11, minutes)
            ws.write(idx, 12, frequency)

    wb.save(response)
    return response


def export_daily_oee_report_xls(request):
    query = request.GET.get("q")
    report_list = Pareto.objects.filter(
        Q(pareto_date__exact=query)
    ).order_by("-user")
    object_list = get_details_to_display(object_list=report_list)

    response = HttpResponse(content_type="application/ms-excel")
    response["Content-Disposition"] = 'attachment; filename="Daily_OEE_Report.xls"'

    wb = xlwt.Workbook(encoding="utf-8")
    ws = wb.add_sheet("Report")

    fnt = Font()
    fnt.height = 300
    fnt.bold = True
    style = XFStyle()
    style.font = fnt

    ws.write(0, 0, f"Daily OEE Report - {object_list[0]['date']}", style)
    row_num = 2

    row_list = ["id", "user", "shift", "availability", "performance", "quality", "oee"]

    headers = ["Pareto Id", "Line/User", "Shift", "Availability", "Performance", "Quality", "OEE"]

    percentage = ["availability", "performance", "quality", "oee"]

    for col_num in range(len(row_list)):
        style1 = xlwt.easyxf("font: bold 1")
        ws.write(1, col_num, headers[col_num], style1)

    style2 = XFStyle()
    style2.num_format_str = "0%"

    style3 = XFStyle()
    style3.num_format_str = "D-MMM-YY"

    for row in object_list:
        for col_num in range(len(row_list)):
            dict_key = row_list[col_num]
            elem = row[dict_key]
            if dict_key == "date":
                ws.write(row_num, col_num, elem, style3)
            elif dict_key in percentage:
                perc_value = elem / 100
                ws.write(row_num, col_num, perc_value, style2)
            else:
                ws.write(row_num, col_num, elem)
        row_num += 1

    wb.save(response)
    return response


def pareto_view(request):
    today_pareto = Pareto.objects.filter(pareto_date=datetime.now()).order_by("-user", "shift")

    report_list = get_details_to_display(object_list=today_pareto)

    return render(request,
                  template_name="gemba/pareto_view.html",
                  context={
                      "report_list": report_list,
                  },
                  )


START_CHOICES = (
    (SHIFT_CHOICES[1][1], "06:00:00"),
    (SHIFT_CHOICES[2][1], "14:00:00"),
    (SHIFT_CHOICES[3][1], "22:00:00"),
)


@login_required
def pareto_create_new(request):
    user = request.user
    line_user_qs = LineUser.objects.filter(user=user)

    if line_user_qs.exists():
        line = line_user_qs[0].line

    pareto_date = datetime.now(timezone.utc).date()
    form = NewPareto(request.POST or None)
    if form.is_valid():
        shift = form.cleaned_data["shift"]
        hours = form.cleaned_data["hours"]
        ops = form.cleaned_data["ops"]
        time_start_qs = LineHourModel.objects.filter(line=line, shift=shift)

        if time_start_qs.exists():
            time_stamp_obj = time_start_qs[0].start
            time_stamp = datetime.strptime(str(time_stamp_obj), "%H:%M:%S")
        else:
            if shift == SHIFT_CHOICES[1][1]:
                time_stamp = datetime.strptime(str(START_CHOICES[0][1]), "%H:%M:%S")
            elif shift == SHIFT_CHOICES[2][1]:
                time_stamp = datetime.strptime(str(START_CHOICES[1][1]), "%H:%M:%S")
            else:
                time_stamp = datetime.strptime(str(START_CHOICES[2][1]), "%H:%M:%S")

        if not line_user_qs.exists():
            line = ""

        Pareto.objects.create(user=user, completed=False, shift=shift, hours=hours, pareto_date=pareto_date,
                              time_stamp=time_stamp, ops=ops, line=line)
        return redirect("gemba_app:pareto-summary")

    return render(
        request,
        template_name="form.html",
        context={"form": form}
    )


class ParetoUpdateView(UpdateView):
    model = Pareto
    template_name = "form.html"
    form_class = ParetoUpdateForm
    success_url = reverse_lazy("gemba_app:pareto-summary")


class ScrapDetailView(DetailView):
    model = ScrapDetail
    template_name = "gemba/scrap_detail_view.html"


class ScrapUpdateView(UpdateView):
    model = ScrapDetail
    fields = ("job", "qty",)
    template_name = "form.html"
    success_url = reverse_lazy("gemba_app:pareto-summary")


class ScrapDeleteView(DeleteView):
    model = ScrapDetail
    template_name = "gemba/delete_scrap.html"
    success_url = reverse_lazy("gemba_app:pareto-summary")


class DowntimeDetailView(DetailView):
    model = DowntimeDetail
    template_name = "gemba/downtime_detail_view.html"


class DowntimeUpdateView(UpdateView):
    model = DowntimeDetail
    fields = ("job", "minutes", "frequency",)
    template_name = "form.html"
    success_url = reverse_lazy("gemba_app:pareto-summary")


class DowntimeDeleteView(DeleteView):
    model = DowntimeDetail
    template_name = "gemba/delete_downtime.html"
    success_url = reverse_lazy("gemba_app:pareto-summary")


class ParetoDetailView(DetailView):
    model = ParetoDetail
    template_name = "gemba/pareto_detail_view.html"


@login_required
def pareto_detail_update(request, pk):
    pareto_detail_obj = ParetoDetail.objects.get(pk=pk)
    form = ParetoDetailUpdateForm(instance=pareto_detail_obj)

    old_job = pareto_detail_obj.job
    target = pareto_detail_obj.job.target
    old_output = pareto_detail_obj.output * -1
    modified = pareto_detail_obj.modified

    if request.method == "POST":
        form = ParetoDetailUpdateForm(request.POST, instance=pareto_detail_obj)
        if form.is_valid():
            job = form.cleaned_data["job"]
            output = form.cleaned_data["output"]
            good = form.cleaned_data["good"]
            scrap = form.cleaned_data["scrap"]

            ParetoDetail.objects.update(job=job, output=output, good=good, scrap=scrap)

            if job == old_job:
                add_output = output + old_output
                tools_update(job=old_job, output=add_output, target=target, modified=modified)
            else:
                # tools update
                tools_update(job=old_job, output=old_output, target=target, modified=modified)
                tools_update(job=job, output=output, target=target, modified=modified)

            return redirect("gemba_app:pareto-summary")

    return render(
        request,
        template_name="form.html",
        context={"form": form}
    )


@login_required
def pareto_detail_delete(request, pk):
    pareto_detail_obj = ParetoDetail.objects.get(pk=pk)

    old_job = pareto_detail_obj.job
    target = pareto_detail_obj.job.target
    old_output = pareto_detail_obj.output * -1
    modified = pareto_detail_obj.modified

    if request.method == "POST":
        pareto_detail_obj.delete()
        # tools update
        tools_update(job=old_job, output=old_output, target=target, modified=modified)
        return redirect("gemba_app:pareto-summary")

    return render(
        request,
        template_name="gemba/delete_detail.html",
        context={"object": pareto_detail_obj}
    )


@login_required
def add_scrap_detail(request, pk):
    scrap = get_object_or_404(ScrapDetail, pk=pk)
    old_qty = scrap.qty
    form = ScrapQuantity(request.POST or None)
    if form.is_valid():
        qty = form.cleaned_data["qty"]
        new_qty = qty + old_qty

        scrap.qty = new_qty
        scrap.save()
        return redirect("gemba_app:pareto-summary")

    return render(
        request,
        template_name="form.html",
        context={"form": form}
    )


@login_required
def add_downtime_time(request, pk):
    downtime = get_object_or_404(DowntimeDetail, pk=pk)
    old_time = downtime.minutes
    form = DowntimeMinutes(request.POST or None)
    if form.is_valid():
        minutes = form.cleaned_data["minutes"]
        new_time = minutes + old_time

        downtime.minutes = new_time
        downtime.save()
        return redirect("gemba_app:pareto-summary")

    return render(
        request,
        template_name="form.html",
        context={"form": form}
    )


def downtime_user_list(request):
    pareto = Pareto.objects.get(user=request.user, completed=False)
    group_qs = DowntimeGroup.objects.filter(user=request.user)
    message_status = ""
    job_otg = pareto.job_otg

    if job_otg is None:
        message_status = "Display"

    if group_qs.exists():
        group = group_qs[0]
        items_list = DowntimeUser.objects.filter(group=group).filter(order__gt=0).order_by("order")
    else:
        items_list = {}

    return render(
        request,
        template_name="gemba/downtime_user_view.html",
        context={
            "pareto": pareto,
            "message_status": message_status,
            "items_list": items_list,
        },
    )


def scrap_user_list(request):
    pareto = Pareto.objects.get(user=request.user, completed=False)
    group_qs = DowntimeGroup.objects.filter(user=request.user)
    message_status = ""
    job_otg = pareto.job_otg

    if job_otg is None:
        message_status = "Display"

    if group_qs.exists():
        group = group_qs[0]
        items_list = ScrapUser.objects.filter(group=group).filter(order__gt=0).order_by("order")
    else:
        items_list = {}

    return render(
        request,
        template_name="gemba/scrap_user_view.html",
        context={
            "pareto": pareto,
            "message_status": message_status,
            "items_list": items_list,
        },
    )


def job_user_list(request):
    group = DowntimeGroup.objects.get(user=request.user)
    job_qs = JobModel2.objects.filter(group=group).order_by("name")

    return render(
        request,
        template_name="gemba/job_user_view.html",
        context={
            "job_list": job_qs,
        },
    )


def select_job(request, pk):
    job = get_object_or_404(JobModel2, pk=pk)
    pareto = Pareto.objects.get(user=request.user, completed=False)
    pareto.job_otg = job
    pareto.save()
    return redirect("gemba_app:pareto-summary")


class DowntimeSearchResultsView(ListView):
    model = DowntimeDetail
    template_name = "gemba/downtimes_search.html"
    paginate_by = 100

    def get_queryset(self):
        query = self.request.GET.get("q")
        page_obj = DowntimeDetail.objects.filter(
            Q(pareto_date__icontains=query) | Q(user__username__icontains=query) |
            Q(downtime__code__icontains=query) | Q(downtime__description__icontains=query) |
            Q(minutes__icontains=query) | Q(job__name__icontains=query) |
            Q(pareto_id__icontains=query)
        ).order_by('-modified')
        return page_obj


class ScrapSearchResultsView(ListView):
    model = ScrapDetail
    template_name = "gemba/scraps_search.html"
    paginate_by = 100

    def get_queryset(self):
        query = self.request.GET.get("q")
        page_obj = ScrapDetail.objects.filter(
            Q(pareto_date__icontains=query) | Q(user__username__icontains=query) |
            Q(scrap__code__icontains=query) | Q(scrap__description__icontains=query) |
            Q(qty__icontains=query) | Q(job__name__icontains=query) |
            Q(pareto_id__icontains=query)
        ).order_by('-modified')
        return page_obj


class ParetoDetailsSearchResultsView(ListView):
    model = ParetoDetail
    template_name = "gemba/pareto_details_search.html"
    paginate_by = 100

    def get_queryset(self):
        query = self.request.GET.get("q")
        page_obj = ParetoDetail.objects.filter(
            Q(pareto_date__icontains=query) | Q(user__username__icontains=query) |
            Q(good__icontains=query) | Q(output__icontains=query) |
            Q(scrap__icontains=query) | Q(job__name__icontains=query) |
            Q(pareto_id__icontains=query) | Q(modified__icontains=query) |
            Q(takt_time__icontains=query)
        ).order_by('-modified')
        return page_obj


# alter to Quarantine id
def quarantine_view(request):
    quarantine_qs = ScrapDetail.objects.filter(scrap=18).order_by("-modified")

    return render(
        request,
        template_name="gemba/quarantine_view.html",
        context={
            "quarantine_qs": quarantine_qs,
        },
    )


def export_scrap_search_csv(request):
    query = request.GET.get("q")
    object_list = ScrapDetail.objects.filter(
        Q(pareto_date__icontains=query) | Q(user__username__icontains=query) |
        Q(scrap__code__icontains=query) | Q(scrap__description__icontains=query) |
        Q(qty__icontains=query) | Q(job__name__icontains=query) |
        Q(pareto_id__icontains=query)
    ).order_by('id')
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="Scrap_reasons.csv"'
    writer = csv.writer(response)
    writer.writerow(["Id", "Date", "User/Line", "Code - Description", "Qty", "Job", "Pareto id"])
    for e in object_list:
        writer.writerow([e.id,
                         e.pareto_date,
                         e.user,
                         e.scrap,
                         e.output,
                         e.job,
                         e.pareto_id,
                         ])
    return response


def export_downtime_search_result_csv(request):
    query = request.GET.get("q")
    object_list = DowntimeDetail.objects.filter(
        Q(pareto_date__icontains=query) | Q(user__username__icontains=query) |
        Q(downtime__code__icontains=query) | Q(downtime__description__icontains=query) |
        Q(minutes__icontains=query) | Q(job__name__icontains=query) |
        Q(pareto_id__icontains=query)
    ).order_by('id')
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="Downtime_reasons.csv"'
    writer = csv.writer(response)
    writer.writerow(["Id", "Date", "User/Line", "Code - Description", "Minutes", "Job", "Pareto id"])
    for e in object_list:
        writer.writerow([e.id,
                         e.pareto_date,
                         e.user,
                         e.downtime,
                         e.minutes,
                         e.job,
                         e.pareto_id,
                         ])
    return response


def export_downtimes_xls(request):
    query = request.GET.get("q")
    object_list = DowntimeDetail.objects.filter(
        Q(pareto_date__icontains=query) | Q(user__username__icontains=query) |
        Q(downtime__code__icontains=query) | Q(downtime__description__icontains=query) |
        Q(minutes__icontains=query) | Q(job__name__icontains=query) |
        Q(pareto_id__icontains=query)
    ).order_by('id')

    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="Downtime_reasons.xls"'

    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('Report')

    row_num = 1
    rows = object_list.values('id', 'pareto_date', 'user__username', 'downtime__code', 'downtime__description',
                              'minutes', 'job__name', 'pareto_id')
    row_list = ['id', 'pareto_date', 'user__username', 'downtime__code', 'downtime__description',
                'minutes', 'job__name', 'pareto_id']
    headers = ["Id", "Pareto date", "Line / User", "Code", "Description", "Minutes", "Job", "Pareto Id"]
    for col_num in range(len(row_list)):
        style = xlwt.easyxf('font: bold 1')
        ws.write(0, col_num, headers[col_num], style)

    for row in rows:
        for col_num in range(len(row_list)):
            elem = row[row_list[col_num]]
            ws.write(row_num, col_num, elem)
        row_num += 1
    wb.save(response)

    return response


class ParetoNSUpdateView(UpdateView):
    model = Pareto
    template_name = "form.html"
    form_class = NotScheduledToRunUpdateForm
    success_url = reverse_lazy("gemba_app:pareto-summary")


class EditorChartView(TemplateView):
    template_name = 'gemba/chart.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["qs"] = Editors.objects.all()
        return context


@login_required
def open_pareto(request, pk):
    pareto = Pareto.objects.get(pk=pk)
    user = request.user

    down_items = pareto.downtimes.all()
    down_items.update(completed=False)
    down_items.update(user=user)
    for item in down_items:
        item.save()

    scrap_items = pareto.scrap.all()
    scrap_items.update(completed=False)
    scrap_items.update(user=user)
    for item in scrap_items:
        item.save()

    details_item = pareto.jobs.all()
    details_item.update(completed=False)
    details_item.update(user=user)
    for item in details_item:
        item.save()

    pareto.user = user
    pareto.completed = False
    pareto.save()

    return redirect("gemba_app:pareto-summary")


import pandas as pd
import openpyxl
from pathlib import Path


@login_required
def export_to_gemba(request):
    query = request.GET.get("q")

    pareto_list = Pareto.objects.filter(
        Q(pareto_date__exact=query)
    )

    col_a = ["" for x in range(536)]
    col_b = ["" for x in range(536)]
    col_c = [0 for x in range(536)]
    col_d = [0 for x in range(536)]
    jobs = []
    takt_times = []
    output_values = []
    good_values = []

    if pareto_list.exists():
        for pareto in pareto_list:
            shift = pareto.shift
            if shift == AM:
                shift_vector = 0
            elif shift == PM:
                shift_vector = 159
            else:
                shift_vector = 318

            pareto_id = pareto.id
            # user to fix!
            user = pareto.user
            ops = pareto.ops
            hours = pareto.hours
            ns = pareto.not_scheduled_to_run
            available_time = int(hours) * 60 - ns
            col_c[61 + shift_vector] = available_time
            col_b[61 + shift_vector] = "Total Available Time"

            for elem in pareto.jobs.all():
                elem_job = elem.job.name
                jobs.append(elem_job)
                elem_job_id = elem.job.id
                job_obj = JobModel2.objects.get(id=elem_job_id)
                target = job_obj.target
                elem_takt_time = round(60 / target, ndigits=4)
                takt_times.append(elem_takt_time)
                elem_output = elem.output
                output_values.append(elem_output)
                # elem_good = elem.good
                # good_values.append(elem_good)
            col_b[62 + shift_vector] = "Product A - Run"
            col_c[62 + shift_vector] = jobs[0]
            col_d[62 + shift_vector] = takt_times[0]
            col_b[63 + shift_vector] = "Total Number of Direct Operators"
            col_c[63 + shift_vector] = ops
            col_b[64 + shift_vector] = "Total Bags made(bags)"
            col_c[64 + shift_vector] = output_values[0]
            if len(jobs) > 1:
                col_b[108 + shift_vector] = "Product B - Run"
                col_c[108 + shift_vector] = jobs[1]
                col_d[108 + shift_vector] = takt_times[1]
                col_b[109 + shift_vector] = "Total Number of Direct Operators"
                col_c[109 + shift_vector] = ops
                col_b[110 + shift_vector] = "Total Bags made(bags)"
                col_c[110 + shift_vector] = output_values[1]

            group = DowntimeGroup.objects.get(user=user)
            down_qs = DowntimeUser.objects.filter(group=group).order_by("gemba")

            down_vector = 0
            for idx, down_obj in enumerate(down_qs):
                down_id = down_obj.downtime.id
                down_item = DowntimeModel.objects.get(id=down_id)
                down_code = down_item.code
                down_name = down_item.description
                pos = shift_vector + down_vector + idx
                col_a[pos] = down_code
                col_b[pos] = down_name

                down_exist_obj = DowntimeDetail.objects.filter(pareto_id=pareto_id).filter(downtime=down_id)
                minutes = 0
                frequency = 0
                if down_exist_obj.exists():
                    for down_obj in down_exist_obj:
                        obj_min = down_obj.minutes
                        obj_freq = down_obj.frequency
                        minutes += obj_min
                        frequency += obj_freq
                col_c[pos] = minutes
                col_d[pos] = frequency

            scrap_qs = ScrapUser.objects.filter(group=group).order_by("gemba")

            offset_a = 65
            offset_b = 46
            for idx, scrap_obj in enumerate(scrap_qs):
                scrap_id = scrap_obj.scrap.id
                scrap_item = ScrapModel.objects.get(id=scrap_id)
                scrap_code = scrap_item.code
                scrap_name = scrap_item.description
                pos = shift_vector + offset_a + idx
                col_a[pos] = scrap_code
                col_a[pos + offset_b] = scrap_code
                col_b[pos] = scrap_name
                col_b[pos + offset_b] = scrap_name

                scraps_exist_qs = ScrapDetail.objects.filter(pareto_id=pareto_id).filter(scrap=scrap_id)
                if scraps_exist_qs.exists():
                    for elem in scraps_exist_qs:
                        job = elem.job.name
                        num = jobs.index(job)
                        qty = elem.qty
                        if num == 0:
                            # product A
                            col_c[pos] = qty
                        else:
                            # product B
                            col_c[pos + offset_b] += qty

    path_to_download_folder = str(os.path.join(Path.home(), "Downloads"))

    file_name = f"{path_to_download_folder}\\{query} - GembaData.xlsx"

    raw_data = {"code": col_a,
                "desc": col_b,
                "min": col_c,
                "freq": col_d}

    df = pd.DataFrame(raw_data, columns=["code", "desc", "min", "freq"])

    df.to_excel(file_name, sheet_name="data")

    return redirect("gemba_app:index")
